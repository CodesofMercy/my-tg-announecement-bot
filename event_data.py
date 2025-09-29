import openpyxl
import os

def get_event_data():
    """Получение данных о мероприятиях из Excel."""
    file_path = 'events.xlsx'
    events = []
    
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]  # Предполагаем, что первая строка — заголовки
        for row in sheet.iter_rows(min_row=2, values_only=True):
            event = dict(zip(headers, row))
            events.append(event)
    else:
        # Создаём файл с заголовками, если его нет
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Название', 'Дата', 'Время', 'Место', 'Описание', 'Цена'])
        workbook.save(file_path)
    
    return events