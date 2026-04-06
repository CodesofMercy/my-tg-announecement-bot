"""
User data persistence — Google Sheets (primary) or local Excel (fallback).
"""
import os
import asyncio
import logging
from sheets_client import get_gsheets_client
import config

logger = logging.getLogger(__name__)


async def save_user_to_gsheets(user_data: dict):
    """Save or update user in Google Sheets. Falls back to Excel if Sheets unavailable."""
    if config.HAS_GOOGLE_SHEETS:
        client = get_gsheets_client()
        if client:
            return await _save_to_sheets(client, config.GOOGLE_SHEET_ID_USERS, user_data)
    # Fallback
    await _save_to_excel(user_data)


async def _save_to_sheets(client, sheet_id: str, user_data: dict):
    """Append user row to the 'Users' sheet."""
    def _do():
        workbook = client.open_by_key(sheet_id)
        sheet = workbook.sheet1
        headers = ["user_id", "first_name", "last_name", "phone_number", "username",
                   "email", "company", "position"]
        if not sheet.row_values(1):
            sheet.append_row(headers)
        sheet.append_row([
            user_data.get("user_id", ""),
            user_data.get("first_name", ""),
            user_data.get("last_name", ""),
            user_data.get("phone_number", ""),
            user_data.get("username", ""),
            user_data.get("email", ""),
            user_data.get("company", ""),
            user_data.get("position", ""),
        ])

    try:
        await asyncio.to_thread(_do)
        logger.info("User saved to Sheets: %s %s (id=%s)",
                    user_data.get("first_name", ""), user_data.get("last_name", ""),
                    user_data.get("user_id", ""))
    except Exception as e:
        logger.error("Error saving user to Sheets: %s", e)
        await _save_to_excel(user_data)


async def _save_to_excel(user_data: dict):
    """Fallback: save to local Excel file."""
    try:
        import openpyxl
        path = os.path.join(os.path.dirname(__file__), "users.xlsx")
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
        ws = wb.active
        if not ws["A1"].value:
            ws.append(["user_id", "first_name", "last_name", "phone_number",
                       "username", "email", "company", "position"])
        ws.append([
            user_data.get("user_id", ""),
            user_data.get("first_name", ""),
            user_data.get("last_name", ""),
            user_data.get("phone_number", ""),
            user_data.get("username", ""),
            user_data.get("email", ""),
            user_data.get("company", ""),
            user_data.get("position", ""),
        ])
        wb.save(path)
        logger.info("User saved to Excel: %s %s", user_data.get("first_name", ""),
                    user_data.get("last_name", ""))
    except Exception as e:
        logger.error("Error saving user to Excel: %s", e)


async def get_user_data(user_id: int) -> dict | None:
    """Look up user data by user_id. Returns dict or None."""
    if config.HAS_GOOGLE_SHEETS:
        client = get_gsheets_client()
        if client:
            result = await _get_from_sheets(client, config.GOOGLE_SHEET_ID_USERS, user_id)
            if result:
                return result
    return await _get_from_excel(user_id)


async def _get_from_sheets(client, sheet_id: str, user_id: int) -> dict | None:
    def _do():
        workbook = client.open_by_key(sheet_id)
        sheet = workbook.sheet1
        values = sheet.get_all_values()
        if not values:
            return None
        headers = values[0]
        for row in values[1:]:
            row_dict = dict(zip(headers, row))
            if row_dict.get("user_id") == str(user_id):
                return row_dict
        return None
    try:
        return await asyncio.to_thread(_do)
    except Exception as e:
        logger.debug("Sheet read failed: %s", e)
        return None


async def _get_from_excel(user_id: int) -> dict | None:
    try:
        import openpyxl
        path = os.path.join(os.path.dirname(__file__), "users.xlsx")
        if not os.path.exists(path):
            return None
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            if str(row_dict.get("user_id")) == str(user_id):
                return row_dict
    except Exception:
        pass
    return None
