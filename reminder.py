"""
Reminders — upcoming events, registered users, notification sending.
"""
import os
import asyncio
import logging
from datetime import datetime, timedelta
from sheets_client import get_gsheets_client
from event_data import get_event_data
import config

logger = logging.getLogger(__name__)


async def get_upcoming_events(days_ahead: int = 60) -> list:
    """Return upcoming events within the next ``days_ahead`` days."""
    events = await get_event_data()
    today = datetime.now().date()
    cutoff = today + timedelta(days=days_ahead)
    result = []
    for ev in events:
        name = ev.get("Название", "").strip()
        date_str = ev.get("Дата", "").strip()
        if not name:
            continue
        result.append({"name": name, "date": date_str})
    return result


async def get_registered_users(event_name: str) -> list:
    """Return list of (user_id, name) for a given event."""
    if config.HAS_GOOGLE_SHEETS:
        client = get_gsheets_client()
        if client:
            return await _get_from_sheets(client, event_name)
    return await _get_from_excel(event_name)


async def _get_from_sheets(client, event_name: str) -> list:
    def _do():
        sheet = client.open_by_key(config.GOOGLE_SHEET_ID_REGISTRATIONS).sheet1
        values = sheet.get_all_values()
        if not values:
            return []
        headers = values[0]
        users = []
        for row in values[1:]:
            row_dict = dict(zip(headers, row))
            if row_dict.get("event", "") == event_name:
                users.append({
                    "user_id": row_dict.get("user_id", ""),
                    "name": row_dict.get("name", ""),
                })
        return users
    try:
        return await asyncio.to_thread(_do)
    except Exception as e:
        logger.error("Sheet read failed: %s", e)
        return []


async def _get_from_excel(event_name: str) -> list:
    try:
        import openpyxl
        path = os.path.join(os.path.dirname(__file__), "registrations.xlsx")
        if not os.path.exists(path):
            return []
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            if row_dict.get("event") == event_name:
                users.append({"user_id": str(row_dict.get("user_id", "")), "name": row_dict.get("name", "")})
        return users
    except Exception:
        return []


async def send_reminder_to_event(user_id: int, bot, event_name: str, message: str):
    """Send a reminder message to a single user."""
    try:
        await bot.send_message(chat_id=int(user_id), text=message)
        logger.info("Reminder sent to user %s for %s", user_id, event_name)
    except Exception as e:
        logger.warning("Failed to send reminder to user %s: %s", user_id, e)
